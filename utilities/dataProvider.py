import csv
import os
import openpyxl
import pandas

from utilities import configReader


def get_data(sheetName):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Directory where configReader.py is located
    EXCEL_PATH = os.path.join(BASE_DIR, "..", "testdata", "testdata.xlsx")  # Navigate to conf.ini

    # workbook = openpyxl.load_workbook("testdata//testdata.xlsx")
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList


def get_data_from_csv(file):
    file_path = configReader.readConfig("data_files", file)
    file = csv.DictReader(open(file_path))
    data = []
    for element in file:
        data.append(element)
    print(data)
    return data


def read_csv_data(file):
    file_path = configReader.readConfig("data_files", file)
    data = []
    with open(file_path, "r") as csv_file:
        reader = csv.reader(csv_file, delimiter=',')
        next(reader)
        for row in reader:
            data.append(tuple(row))
    return data


