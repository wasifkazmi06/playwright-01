import os

import openpyxl


def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Directory where configReader.py is located
path = os.path.join(BASE_DIR, "..", "testdata", "testdata.xlsx")  # Navigate to conf.ini

# path = "..//testdata//testdata.xlsx"
sheetName = "LoginTest"

rows = getRowCount(path, sheetName)
cols = getColCount(path, sheetName)

print(rows, "---", cols)

print(getCellData(path, sheetName, 2, 1))
setCellData(path, sheetName, 1, 4, "DOB")
